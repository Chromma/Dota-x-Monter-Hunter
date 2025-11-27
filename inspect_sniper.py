#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Inspecciona la estructura completa de la hoja Sniper para entender el layout.
"""
import openpyxl

wb = openpyxl.load_workbook('Dota_MonHunter.xlsx')
ws = wb['Sniper']

print(f"=== Sniper Sheet ===")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}\n")

# Leer los encabezados (fila 1)
print("HEADERS (Row 1):")
headers = []
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(1, col_idx).value
    headers.append(val if val else f"Col{col_idx}")
    if val:
        print(f"  Col {col_idx}: {val}")

print("\n" + "="*100)
print("DATOS (Todas las filas):\n")

for row_idx in range(2, ws.max_row + 1):
    row_data = []
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row_idx, col_idx).value
        row_data.append(val)
    
    # Mostrar en formato legible
    first_col = row_data[0] if row_data[0] else ""
    print(f"Row {row_idx:2d}: {str(first_col):20s} | ", end="")
    
    for col_idx in range(1, min(15, len(row_data))):
        val = row_data[col_idx]
        if val:
            print(f"{val:5} ", end="")
        else:
            print(f"{'':5} ", end="")
    print()
