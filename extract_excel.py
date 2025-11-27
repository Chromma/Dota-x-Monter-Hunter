#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extrae datos del Excel Dota_MonHunter.xlsx de forma limpia y simple.
Salida: excel_data.json con estructura clara por héroe/skin/pieza.
"""
import json
import openpyxl
from collections import defaultdict

# Cargar Excel
wb = openpyxl.load_workbook('Dota_MonHunter.xlsx')

# Mapeo de nombre de héroe Excel -> nombre usado internamente
HERO_SHEETS = {
    'Sniper': 'Sniper',
    'Techies': 'Techies',
    'Anti Mage': 'Anti Mage',
    'Dragon Knight': 'Dragon Knight',
    'Beastmaster': 'Beastmaster',
    'Windranger': 'Windranger',
}

# Mapeo de nombres de materiales Excel -> MAT_IDS keys en constants.js
MATERIAL_MAP = {
    'Monster Essense': 'Essence',
    'Monster Essence': 'Essence',
    'Monster Fur': 'Fur',
    'Monster Claw': 'Claw',
    'Odogaron Sinew': 'OdoSinew',
    'Odogaron Hardfang': 'OdoFang',
    'Odogaron Scale': 'OdoScale',
    'Odogaron Shard': 'OdoShard',
    'Odogaron Mantle': 'OdoMantle',
    'Zinogre Electrofur': 'ZinFur',
    'Zinogre Deathly Shocker': 'ZinShocker',
    'Zinogre Cortex': 'ZinCortex',
    'Zinogre Hardhorn': 'ZinHorn',
    'Zinogre Skymerald': 'ZinSky',
    'Rathalos Carapache': 'RathCarapache',
    'Rathalos Wing': 'RathWing',
    'Rathalos Tail': 'RathTail',
    'Rathalos Ruby': 'RathRuby',
    'Rathalos Plate': 'RathPlate',
    'Kirin Thunderhorn': 'KirThunderhorn',
    'Kirin Hide': 'KirHide',
    'Kirin Mane': 'KirMane',
    'Kirin Azure Horn': 'KirAzureHorn',
    'Ancient Bone': 'AncientBone',
    'Kestodon Shell': 'Kestshell',
    'Bullfango Head': 'BullHead',
    'Large Barrel': 'LargeBarrel',
    'Gunpowder': 'Gunpowder',
    'Iron Ore': 'IronOre',
    "Devil's Blight": 'DevilsBlight',
    'Sinister Cloth': 'SinisterCloth',
    'Large Elder Dragon Gem': 'ElderDragGem',
    'Wyvern Gem': 'WyvernGem',
    'Warped Bone': 'WarpedBone',
    'Prismatic Pigment': 'PrismaticPigment',
}

result = {}

for sheet_name, hero_name in HERO_SHEETS.items():
    if sheet_name not in wb.sheetnames:
        print(f"WARN: Sheet '{sheet_name}' not found")
        continue
    
    ws = wb[sheet_name]
    print(f"\nProcessing sheet: {sheet_name}")
    
    # Leer encabezados (primera fila)
    headers = []
    for col in range(1, ws.max_column + 1):
        h = ws.cell(1, col).value
        headers.append(h if h else f'Col{col}')
    
    hero_data = {}
    current_skin = None
    current_pieces = {}
    
    for row_idx in range(2, ws.max_row + 1):
        row_dict = {}
        for col_idx, header in enumerate(headers, 1):
            row_dict[header] = ws.cell(row_idx, col_idx).value
        
        # Columna que identifica skin
        skin_name = row_dict.get(hero_name)
        piece_name = row_dict.get('Unnamed: 1')
        
        # Si hay nombre de skin, cambiamos de skin
        if skin_name and skin_name.strip().lower() != 'total':
            if current_skin and current_pieces:
                hero_data[current_skin] = dict(current_pieces)
            current_skin = skin_name
            current_pieces = {}
            continue
        
        # Si hay nombre de pieza y no es "Total", es una pieza de recurso
        if piece_name and piece_name.strip().lower() != 'total' and current_skin:
            if piece_name not in current_pieces:
                current_pieces[piece_name] = []
            
            # Extraer materiales y cantidades
            for header, value in row_dict.items():
                if header in [hero_name, 'Unnamed: 1', None]:
                    continue
                if isinstance(value, (int, float)) and value > 0 and header in MATERIAL_MAP:
                    current_pieces[piece_name].append({
                        'material': MATERIAL_MAP[header],
                        'amount': int(value)
                    })
    
    # Guardar última skin
    if current_skin and current_pieces:
        hero_data[current_skin] = dict(current_pieces)
    
    result[hero_name] = hero_data
    print(f"  Skins found: {list(hero_data.keys())}")

# Guardar resultado
with open('excel_data.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, indent=2, ensure_ascii=False)

print(f"\n✓ Extracted data saved to excel_data.json")
