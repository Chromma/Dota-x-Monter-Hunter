#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extrae datos del Excel: Hero > Set > Piece > Materials
Estructura: Col1=SetName, Col2=PieceName, Col3+=Materials
"""
import openpyxl
import json

# Mapeo EXACTO de materiales Excel -> MAT_IDS
MATERIAL_MAP = {
    'Monster Essense': 'Essence',
    'Monster Fur': 'Fur',
    'Monster Claw': 'Claw',
    'Odogaron Sinew': 'OdoSinew',
    'Odogaron Hardfang': 'OdoFang',
    'Odogaron Scale': 'OdoScale',
    'Zinogre Electrofur': 'ZinFur',
    'Zinogre Deathly Shocker': 'ZinShocker',
    'Zinogre Cortex': 'ZinCortex',
    'Rathalos Carapache': 'RathCarapache',
    'Rathalos Wing': 'RathWing',
    'Rathalos Tail': 'RathTail',
    'Kirin Thunderhorn': 'KirThunderhorn',
    'Kirin Hide': 'KirHide',
    'Kirin Mane': 'KirMane',
    'Ancient Bone': 'AncientBone',
    'Kestodon Shell': 'KestShell',
    'Bullfango Head': 'BullHead',
    'Large Barrel': 'LargeBarrel',
    'Gunpowder': 'Gunpowder',
    'Iron Ore': 'IronOre',
    'Odogaron Shard': 'OdoShard',
    'Odogaron Mantle': 'OdoMantle',
    'Zinogre Hardhorn': 'ZinHorn',
    'Zinogre Skymerald': 'ZinSky',
    'Rathalos Ruby': 'RathRuby',
    'Rathalos Plate': 'RathPlate',
    'Kirin Azure Horn': 'KirAzureHorn',
    'Large Elder Dragon Gem': 'ElderDragGem',
    'Wyvern Gem': 'WyvernGem',
    'Warped Bone': 'WarpedBone',
    'Sinister Cloth': 'SinisterCloth',
    "Devil's Blight": 'DevilsBlight',
    'Prismatic Pigment': 'PrismaticPigment',
}

HEROES = ['Sniper', 'Techies', 'Anti Mage', 'Dragon Knight', 'Beastmaster', 'Windranger']

wb = openpyxl.load_workbook('Dota_MonHunter.xlsx')
result = {}

for hero_name in HEROES:
    if hero_name not in wb.sheetnames:
        print(f"⚠ Sheet '{hero_name}' not found")
        continue
    
    ws = wb[hero_name]
    print(f"\n{'='*120}")
    print(f"HERO: {hero_name.upper()}")
    print(f"{'='*120}")
    
    # Row 2 contiene encabezados de materiales (a partir de Col3)
    material_headers = []
    for col_idx in range(3, ws.max_column + 1):
        header = ws.cell(2, col_idx).value
        if header and header.strip().lower() not in ['total', '']:
            material_headers.append((col_idx, header))
    
    hero_data = {}
    current_set = None
    
    # Recorrer desde row 3 en adelante
    row_idx = 3
    while row_idx <= ws.max_row:
        col1_value = ws.cell(row_idx, 1).value
        col2_value = ws.cell(row_idx, 2).value
        
        # Detectar Set: Col1 tiene valor y no es vacío (primera row del set)
        if col1_value and col1_value.strip() and col1_value.strip().lower() != 'total':
            current_set = col1_value.strip()
            if current_set not in hero_data:
                hero_data[current_set] = {}
            print(f"\n  SET: {current_set}")
        
        # Detectar Pieza: Col2 tiene valor y no es "Total"
        if col2_value and col2_value.strip() and col2_value.strip().lower() != 'total' and current_set:
            piece_name = col2_value.strip()
            piece_data = {}
            
            # Extraer materiales de esta pieza (desde Col3 en adelante)
            for col_idx, material_name in material_headers:
                amount = ws.cell(row_idx, col_idx).value
                if amount and isinstance(amount, (int, float)) and amount > 0:
                    mat_key = MATERIAL_MAP.get(material_name)
                    if mat_key:
                        piece_data[mat_key] = int(amount)
            
            if piece_data:
                hero_data[current_set][piece_name] = piece_data
                total = sum(piece_data.values())
                mat_list = ", ".join([f"{k}:{v}" for k, v in sorted(piece_data.items())])
                print(f"    {piece_name:15s} | Total: {total:3d} | {mat_list}")
        
        row_idx += 1
    
    result[hero_name] = hero_data

# Guardar JSON
with open('excel_data.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, indent=2, ensure_ascii=False)

print(f"\n{'='*120}")
print("✓ Datos extraídos guardados en excel_data.json")
print(f"{'='*120}\n")

# Resumen
print("RESUMEN DE EXTRACCIÓN:\n")
for hero, sets in result.items():
    print(f"{hero}:")
    for set_name, pieces in sets.items():
        total_resources = sum(sum(p.values()) for p in pieces.values())
        print(f"  {set_name:15s}: {len(pieces)} piezas | Total recursos: {total_resources}")
