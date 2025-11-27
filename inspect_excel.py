#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Inspecciona la estructura del Excel para entender cómo está organizado.
"""
import openpyxl

wb = openpyxl.load_workbook('Dota_MonHunter.xlsx')
print("Sheet names:", wb.sheetnames)

for sheet_name in wb.sheetnames[:1]:  # Solo Sniper
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    print("\nFirst 20 rows:")
    for row_idx in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(10, ws.max_column + 1)):
            val = ws.cell(row_idx, col_idx).value
            row_data.append(str(val)[:15] if val else "")
        print(f"Row {row_idx}: {row_data}")
